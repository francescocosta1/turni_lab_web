import calendar as pycalendar
from collections import defaultdict
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import AssenzaForm, CalendarioMensileForm, DipendenteForm
from .models import Assenza, AssegnazioneTurno, CalendarioMensile, Dipendente
from .services import genera_turni_mese,ripianifica_calendario


def fabbisogno_giornaliero(data_corrente):
    """
    Caso pseudo-reale:
    - lun-ven: M 7, P 5, N 1
    - sabato:  M 5, P 4, N 1
    - domenica: M 4, P 3, N 1
    """
    giorno_settimana = data_corrente.weekday()

    if giorno_settimana <= 4:
        return {"M": 7, "P": 5, "N": 1}
    elif giorno_settimana == 5:
        return {"M": 5, "P": 4, "N": 1}
    else:
        return {"M": 4, "P": 3, "N": 1}


def utente_admin(user):
    return user.is_authenticated and (
        user.is_superuser or user.groups.filter(name="Amministratori").exists()
    )


def utente_dipendente(user):
    return user.is_authenticated and user.groups.filter(name="Dipendenti").exists()


def valida_modifica_turno(assegnazione, nuovo_turno):
    errori = []

    dipendente = assegnazione.dipendente
    calendario = assegnazione.calendario
    data_corrente = assegnazione.data

    assenza = Assenza.objects.filter(
        dipendente=dipendente,
        data_inizio__lte=data_corrente,
        data_fine__gte=data_corrente
    ).exists()

    if assenza and nuovo_turno in ["M", "P", "N", "R"]:
        errori.append(
            "Il dipendente risulta assente in questa data e non può essere assegnato a questo turno."
        )

    giorno_precedente = data_corrente - timedelta(days=1)
    giorno_successivo = data_corrente + timedelta(days=1)

    assegnazione_precedente = AssegnazioneTurno.objects.filter(
        calendario=calendario,
        dipendente=dipendente,
        data=giorno_precedente
    ).first()

    assegnazione_successiva = AssegnazioneTurno.objects.filter(
        calendario=calendario,
        dipendente=dipendente,
        data=giorno_successivo
    ).first()

    if assegnazione_precedente and assegnazione_precedente.turno == "N":
        if nuovo_turno != "R":
            errori.append("Dopo una notte, il giorno successivo deve essere riposo.")

    if nuovo_turno == "N":
        if assegnazione_successiva and assegnazione_successiva.turno != "R":
            errori.append("Se assegni una notte, il giorno successivo del dipendente deve essere riposo.")

    assegnazioni_giornaliere = AssegnazioneTurno.objects.filter(
        calendario=calendario,
        data=data_corrente
    )

    conteggi = {"M": 0, "P": 0, "N": 0}

    for a in assegnazioni_giornaliere:
        turno = nuovo_turno if a.id == assegnazione.id else a.turno
        if turno in conteggi:
            conteggi[turno] += 1

    fabbisogno = fabbisogno_giornaliero(data_corrente)

    if conteggi["M"] != fabbisogno["M"]:
        errori.append(
            f"Dopo la modifica, il turno mattina del {data_corrente.strftime('%d/%m/%Y')} "
            f"sarebbe {conteggi['M']}/{fabbisogno['M']}."
        )
    if conteggi["P"] != fabbisogno["P"]:
        errori.append(
            f"Dopo la modifica, il turno pomeriggio del {data_corrente.strftime('%d/%m/%Y')} "
            f"sarebbe {conteggi['P']}/{fabbisogno['P']}."
        )
    if conteggi["N"] != fabbisogno["N"]:
        errori.append(
            f"Dopo la modifica, il turno notte del {data_corrente.strftime('%d/%m/%Y')} "
            f"sarebbe {conteggi['N']}/{fabbisogno['N']}."
        )

    return errori


def home(request):
    return render(request, "scheduler/home.html")


def login_view(request):
    if request.user.is_authenticated:
        return redirect("smista_dashboard")

    next_url = request.GET.get("next") or request.POST.get("next") or ""

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if next_url:
                return redirect(next_url)
            return redirect("smista_dashboard")
        else:
            messages.error(request, "Credenziali non valide.")

    return render(request, "scheduler/login.html", {"next": next_url})


def logout_view(request):
    logout(request)
    messages.success(request, "Logout effettuato correttamente.")
    return redirect("home")


@login_required
def smista_dashboard(request):
    if utente_admin(request.user):
        return redirect("admin_dashboard")
    if utente_dipendente(request.user):
        return redirect("dipendente_dashboard")

    messages.error(request, "Il tuo account non è associato a un ruolo valido.")
    return redirect("logout_view")


@login_required
def admin_dashboard(request):
    if not utente_admin(request.user):
        messages.error(request, "Non hai i permessi per accedere all'area amministrativa.")
        return redirect("smista_dashboard")

    return render(request, "scheduler/admin_dashboard.html")


@login_required
def dipendente_dashboard(request):
    if not (utente_admin(request.user) or utente_dipendente(request.user)):
        messages.error(request, "Non hai i permessi per accedere all'area dipendente.")
        return redirect("logout_view")

    return render(request, "scheduler/dipendente_dashboard.html")


@login_required
def dipendenti_lista(request):
    if not utente_admin(request.user):
        return redirect("home")

    query = request.GET.get("q")
    dipendenti = Dipendente.objects.all()

    if query:
        dipendenti = dipendenti.filter(
            Q(nome__icontains=query) |
            Q(cognome__icontains=query)
        )

    totale = Dipendente.objects.count()
    senior = Dipendente.objects.filter(livello="senior").count()
    junior = Dipendente.objects.filter(livello="junior").count()
    attivi = Dipendente.objects.filter(attivo=True).count()

    return render(
        request,
        "scheduler/dipendenti_lista.html",
        {
            "dipendenti": dipendenti,
            "query": query,
            "totale": totale,
            "senior": senior,
            "junior": junior,
            "attivi": attivi,
        }
    )


@login_required
def dipendente_nuovo(request):
    if not utente_admin(request.user):
        return redirect("home")

    if request.method == "POST":
        form = DipendenteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Dipendente inserito correttamente.")
            return redirect("dipendenti_lista")
    else:
        form = DipendenteForm()

    return render(
        request,
        "scheduler/dipendente_form.html",
        {
            "form": form,
            "titolo": "Nuovo dipendente",
        }
    )


@login_required
def dipendente_modifica(request, dipendente_id):
    if not utente_admin(request.user):
        return redirect("home")

    dipendente = get_object_or_404(Dipendente, id=dipendente_id)

    if request.method == "POST":
        form = DipendenteForm(request.POST, instance=dipendente)
        if form.is_valid():
            form.save()
            messages.success(request, "Dipendente aggiornato.")
            return redirect("dipendenti_lista")
    else:
        form = DipendenteForm(instance=dipendente)

    return render(
        request,
        "scheduler/dipendente_form.html",
        {
            "form": form,
            "titolo": "Modifica dipendente",
        }
    )


@login_required
def dipendente_elimina(request, dipendente_id):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    dipendente = get_object_or_404(Dipendente, id=dipendente_id)

    if request.method == "POST":
        dipendente.delete()
        messages.success(request, "Dipendente eliminato correttamente.")
        return redirect("dipendenti_lista")

    return render(
        request,
        "scheduler/conferma_elimina.html",
        {
            "oggetto": dipendente,
            "titolo": "Elimina dipendente",
            "messaggio": "Sei sicuro di voler eliminare questo dipendente?",
            "url_annulla": "dipendenti_lista",
        }
    )


@login_required
def calendari_lista(request):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    calendari = CalendarioMensile.objects.all().order_by("-anno", "-mese")
    return render(
        request,
        "scheduler/calendari_lista.html",
        {"calendari": calendari}
    )


@login_required
def calendario_nuovo(request):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    if request.method == "POST":
        form = CalendarioMensileForm(request.POST)
        if form.is_valid():
            mese = form.cleaned_data["mese"]
            anno = form.cleaned_data["anno"]

            if CalendarioMensile.objects.filter(mese=mese, anno=anno).exists():
                messages.error(request, "Esiste già un calendario per questo mese.")
            else:
                form.save()
                messages.success(request, "Calendario mensile creato correttamente.")
                return redirect("calendari_lista")
    else:
        form = CalendarioMensileForm()

    return render(
        request,
        "scheduler/calendario_form.html",
        {"form": form}
    )


@login_required
def calendario_elimina(request, calendario_id):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    calendario = get_object_or_404(CalendarioMensile, id=calendario_id)

    if request.method == "POST":
        calendario.delete()
        messages.success(request, "Calendario eliminato correttamente.")
        return redirect("calendari_lista")

    return render(
        request,
        "scheduler/conferma_elimina.html",
        {
            "oggetto": calendario,
            "titolo": "Elimina calendario",
            "messaggio": "Sei sicuro di voler eliminare questo calendario mensile?",
            "url_annulla": "calendari_lista",
        }
    )


@login_required
def genera_turni_view(request, calendario_id):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    calendario = get_object_or_404(CalendarioMensile, id=calendario_id)

    try:
        genera_turni_mese(calendario)
        messages.success(request, f"I turni del mese {calendario} sono stati generati correttamente.")
    except Exception as e:
        messages.error(request, f"Errore nella generazione dei turni: {e}")

    return redirect("calendario_dettaglio", calendario_id=calendario.id)

@login_required
def ripianifica_calendario_view(request, calendario_id):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    if request.method != "POST":
        return redirect("calendario_dettaglio", calendario_id=calendario_id)

    calendario = get_object_or_404(CalendarioMensile, id=calendario_id)

    try:
        esito = ripianifica_calendario(calendario)

        if esito["assenze_aggiornate"] == 0:
            messages.info(request, "Nessuna nuova assenza da ripianificare.")
        else:
            messages.success(
                request,
                f"Ripianificazione completata: "
                f"{esito['assenze_aggiornate']} assenze aggiornate, "
                f"{esito['coperture_sistemate']} coperture sistemate,"
                f"{esito.get('scambi_effettuati',0)} scambi locali effettuati."
            )
    except Exception as e:
        messages.error(request, f"Ripianificazione non riuscita: {e}")

    return redirect("calendario_dettaglio", calendario_id=calendario.id)

@login_required
def calendario_dettaglio(request, calendario_id):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    calendario = get_object_or_404(CalendarioMensile, id=calendario_id)
    dipendenti = list(Dipendente.objects.filter(attivo=True).order_by("cognome", "nome"))

    numero_giorni = pycalendar.monthrange(calendario.anno, calendario.mese)[1]
    giorni = []
    giorni_settimana_it = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]

    for g in range(1, numero_giorni + 1):
        data_corrente = date(calendario.anno, calendario.mese, g)
        giorni.append({
            "numero": g,
            "sigla": giorni_settimana_it[data_corrente.weekday()],
            "weekend": data_corrente.weekday() >= 5,
        })

    assegnazioni = AssegnazioneTurno.objects.filter(
        calendario=calendario
    ).select_related("dipendente")

    mappa = defaultdict(dict)
    riepiloghi = {
        d.id: {
            "M": 0, "P": 0, "N": 0, "R": 0,
            "F": 0, "A": 0, "L": 0, "X": 0,
            "totale": 0,
            "weekend_lavorati": 0,
        }
        for d in dipendenti
    }

    coperture = {
        g["numero"]: {"M": 0, "P": 0, "N": 0, "R": 0, "F": 0, "A": 0, "L": 0, "X": 0}
        for g in giorni
    }

    for a in assegnazioni:
        giorno = a.data.day
        mappa[a.dipendente_id][giorno] = a
        riepiloghi[a.dipendente_id][a.turno] += 1
        coperture[giorno][a.turno] += 1

        if a.turno in ["M", "P", "N"]:
            riepiloghi[a.dipendente_id]["totale"] += 1
            if a.data.weekday() >= 5:
                riepiloghi[a.dipendente_id]["weekend_lavorati"] += 1

    full_time_ids = [d.id for d in dipendenti if d.tipo_contratto == "full_time"]
    part_time_ids = [d.id for d in dipendenti if d.tipo_contratto == "part_time"]

    media_full_time = 0
    media_part_time = 0

    if full_time_ids:
        media_full_time = sum(riepiloghi[d_id]["totale"] for d_id in full_time_ids) / len(full_time_ids)

    if part_time_ids:
        media_part_time = sum(riepiloghi[d_id]["totale"] for d_id in part_time_ids) / len(part_time_ids)

    media_notti_full_time = 0
    if full_time_ids:
        media_notti_full_time = (
            sum(riepiloghi[d_id]["N"] for d_id in full_time_ids) / len(full_time_ids)
        )

# Separo la media_weekend tra full-time e part_time

    media_weekend_full_time = 0
    media_weekend_part_time = 0

    if full_time_ids:
        media_weekend_full_time = (
            sum(riepiloghi[d_id]["weekend_lavorati"] for d_id in full_time_ids) / len(full_time_ids)
        )

    if part_time_ids:
        media_weekend_part_time = (
            sum(riepiloghi[d_id]["weekend_lavorati"] for d_id in part_time_ids) / len(part_time_ids)
        )


    righe = []
    for d in dipendenti:
        turni_giornalieri = []

        for g in giorni:
            assegnazione = mappa[d.id].get(g["numero"])
            turno = assegnazione.turno if assegnazione else ""

            turni_giornalieri.append({
                "valore": turno,
                "id": assegnazione.id if assegnazione else None,
                "weekend": g["weekend"],
            })

        totale = riepiloghi[d.id]["totale"]
        notti = riepiloghi[d.id]["N"]
        weekend_lavorati = riepiloghi[d.id]["weekend_lavorati"]

        media_riferimento = media_full_time if d.tipo_contratto == "full_time" else media_part_time

        if totale > media_riferimento + 1:
            stato_totale = "alto"
        elif totale < media_riferimento - 1:
            stato_totale = "basso"
        else:
            stato_totale = "ok"

        if d.tipo_contratto == "part_time":
            stato_notti = "non_previsto"
        else:
            if notti > media_notti_full_time + 1:
                stato_notti = "alto"
            elif notti < media_notti_full_time - 1:
                stato_notti = "basso"
            else:
                stato_notti = "ok"

        media_weekend_riferimento = (
            media_weekend_full_time if d.tipo_contratto == "full_time"
            else media_weekend_part_time
        )

        if weekend_lavorati > media_weekend_riferimento + 1:
            stato_weekend = "alto"
        elif weekend_lavorati < media_weekend_riferimento - 1:
            stato_weekend = "basso"
        else:
            stato_weekend = "ok"


        righe.append({
            "dipendente": d,
            "turni": turni_giornalieri,
            "riepilogo": riepiloghi[d.id],
            "equita_totale": stato_totale,
            "equita_notti": stato_notti,
            "equita_weekend": stato_weekend,
        })

    coperture_righe = []

    for label, sigla in [
        ("Mattina", "M"),
        ("Pomeriggio", "P"),
        ("Notte", "N"),
    ]:
        valori = []

        for g in giorni:
            data_corrente = date(calendario.anno, calendario.mese, g["numero"])
            richiesto = fabbisogno_giornaliero(data_corrente)[sigla]
            conteggio = coperture[g["numero"]][sigla]

            valori.append({
                "conteggio": conteggio,
                "richiesto": richiesto,
                "ok": conteggio == richiesto,
                "weekend": g["weekend"],
            })

        coperture_righe.append({
            "label": label,
            "sigla": sigla,
            "valori": valori,
        })

    return render(
        request,
        "scheduler/calendario_dettaglio.html",
        {
            "calendario": calendario,
            "giorni": giorni,
            "righe": righe,
            "coperture_righe": coperture_righe,
            "media_full_time": round(media_full_time, 1),
            "media_part_time": round(media_part_time, 1),
            "media_notti": round(media_notti_full_time, 1),
            "media_weekend_full_time": round(media_weekend_full_time, 1),
            "media_weekend_part_time": round(media_weekend_part_time, 1)
        }
    )


@login_required
def esporta_calendario_excel(request, calendario_id):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    calendario = get_object_or_404(CalendarioMensile, id=calendario_id)
    dipendenti = list(Dipendente.objects.filter(attivo=True).order_by("cognome", "nome"))
    assegnazioni = AssegnazioneTurno.objects.filter(calendario=calendario).select_related("dipendente")

    numero_giorni = pycalendar.monthrange(calendario.anno, calendario.mese)[1]

    mappa = defaultdict(dict)
    riepiloghi = {
        d.id: {"M": 0, "P": 0, "N": 0, "R": 0, "F": 0, "A": 0, "L": 0, "X": 0, "totale": 0}
        for d in dipendenti
    }

    for a in assegnazioni:
        giorno = a.data.day
        mappa[a.dipendente_id][giorno] = a.turno
        riepiloghi[a.dipendente_id][a.turno] += 1
        if a.turno in ["M", "P", "N"]:
            riepiloghi[a.dipendente_id]["totale"] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendario.mese:02d}-{calendario.anno}"

    titolo = f"Turnistica laboratorio - {calendario.mese:02d}/{calendario.anno}"
    ws.cell(row=1, column=1, value=titolo)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    header_row = 3
    ws.cell(row=header_row, column=1, value="Dipendente")

    for giorno in range(1, numero_giorni + 1):
        ws.cell(row=header_row, column=1 + giorno, value=giorno)

    col_base = 2 + numero_giorni
    intestazioni_finali = ["M", "P", "N", "R", "F", "A", "L", "X", "Tot"]

    for i, label in enumerate(intestazioni_finali):
        ws.cell(row=header_row, column=col_base + i, value=label)

    fill_header = PatternFill("solid", fgColor="D9EAF7")
    fill_m = PatternFill("solid", fgColor="DBEAFE")
    fill_p = PatternFill("solid", fgColor="FEF3C7")
    fill_n = PatternFill("solid", fgColor="FEE2E2")
    fill_r = PatternFill("solid", fgColor="E5E7EB")
    fill_f = PatternFill("solid", fgColor="DCFCE7")
    fill_a = PatternFill("solid", fgColor="E0E7FF")
    fill_l = PatternFill("solid", fgColor="FCE7F3")
    fill_x = PatternFill("solid", fgColor="EDE9FE")

    for col in range(1, col_base + len(intestazioni_finali)):
        ws.cell(row=header_row, column=col).fill = fill_header
        ws.cell(row=header_row, column=col).font = Font(bold=True)
        ws.cell(row=header_row, column=col).alignment = Alignment(horizontal="center")

    row = header_row + 1
    for d in dipendenti:
        ws.cell(row=row, column=1, value=f"{d.cognome} {d.nome}")

        for giorno in range(1, numero_giorni + 1):
            turno = mappa[d.id].get(giorno, "")
            cella = ws.cell(row=row, column=1 + giorno, value=turno)
            cella.alignment = Alignment(horizontal="center")

            if turno == "M":
                cella.fill = fill_m
            elif turno == "P":
                cella.fill = fill_p
            elif turno == "N":
                cella.fill = fill_n
            elif turno == "R":
                cella.fill = fill_r
            elif turno == "F":
                cella.fill = fill_f
            elif turno == "A":
                cella.fill = fill_a
            elif turno == "L":
                cella.fill = fill_l
            elif turno == "X":
                cella.fill = fill_x

        ws.cell(row=row, column=col_base + 0, value=riepiloghi[d.id]["M"])
        ws.cell(row=row, column=col_base + 1, value=riepiloghi[d.id]["P"])
        ws.cell(row=row, column=col_base + 2, value=riepiloghi[d.id]["N"])
        ws.cell(row=row, column=col_base + 3, value=riepiloghi[d.id]["R"])
        ws.cell(row=row, column=col_base + 4, value=riepiloghi[d.id]["F"])
        ws.cell(row=row, column=col_base + 5, value=riepiloghi[d.id]["A"])
        ws.cell(row=row, column=col_base + 6, value=riepiloghi[d.id]["L"])
        ws.cell(row=row, column=col_base + 7, value=riepiloghi[d.id]["X"])
        ws.cell(row=row, column=col_base + 8, value=riepiloghi[d.id]["totale"])

        row += 1

    ws.column_dimensions["A"].width = 24

    for col in range(2, 2 + numero_giorni):
        lettera = get_column_letter(col)
        ws.column_dimensions[lettera].width = 5

    for col in range(col_base, col_base + len(intestazioni_finali)):
        lettera = get_column_letter(col)
        ws.column_dimensions[lettera].width = 6

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="turnistica_{calendario.mese:02d}_{calendario.anno}.xlsx"'
    )

    wb.save(response)
    return response


@login_required
def modifica_turno(request, assegnazione_id):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    assegnazione = AssegnazioneTurno.objects.select_related(
        "dipendente", "calendario"
    ).get(id=assegnazione_id)

    turni = [
        ("M", "Mattina"),
        ("P", "Pomeriggio"),
        ("N", "Notte"),
        ("R", "Riposo"),
        ("F", "Ferie"),
        ("L", "Malattia"),
        ("A", "Assenza"),
        ("X", "Permesso"),
    ]

    if request.method == "POST":
        nuovo_turno = request.POST.get("turno")

        if nuovo_turno:
            errori = valida_modifica_turno(assegnazione, nuovo_turno)

            if errori:
                for errore in errori:
                    messages.error(request, errore)
            else:
                assegnazione.turno = nuovo_turno
                assegnazione.save()
                messages.success(request, "Turno aggiornato correttamente.")
                return redirect("calendario_dettaglio", calendario_id=assegnazione.calendario.id)

    return render(
        request,
        "scheduler/modifica_turno.html",
        {
            "assegnazione": assegnazione,
            "turni": turni,
        }
    )


@login_required
def assenze_lista(request):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    assenze = Assenza.objects.select_related("dipendente").all()
    return render(
        request,
        "scheduler/assenze_lista.html",
        {"assenze": assenze}
    )


@login_required
def assenza_nuova(request):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    if request.method == "POST":
        form = AssenzaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Assenza inserita correttamente.")
            return redirect("assenze_lista")
        else:
            messages.error(request, "Controlla i dati inseriti nel modulo.")
    else:
        form = AssenzaForm()

    return render(
        request,
        "scheduler/assenza_form.html",
        {
            "form": form,
            "titolo": "Nuova assenza",
        }
    )


@login_required
def assenza_modifica(request, assenza_id):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    assenza = get_object_or_404(Assenza, id=assenza_id)

    if request.method == "POST":
        form = AssenzaForm(request.POST, instance=assenza)
        if form.is_valid():
            form.save()
            messages.success(request, "Assenza aggiornata correttamente.")
            return redirect("assenze_lista")
        else:
            messages.error(request, "Controlla i dati inseriti nel modulo.")
    else:
        form = AssenzaForm(instance=assenza)

    return render(
        request,
        "scheduler/assenza_form.html",
        {
            "form": form,
            "titolo": "Modifica assenza",
        }
    )


@login_required
def assenza_elimina(request, assenza_id):
    if not utente_admin(request.user):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("smista_dashboard")

    assenza = get_object_or_404(Assenza, id=assenza_id)

    if request.method == "POST":
        assenza.delete()
        messages.success(request, "Assenza eliminata correttamente.")
        return redirect("assenze_lista")

    return render(
        request,
        "scheduler/conferma_elimina.html",
        {
            "oggetto": assenza,
            "titolo": "Elimina assenza",
            "messaggio": "Sei sicuro di voler eliminare questa assenza?",
            "url_annulla": "assenze_lista",
        }
    )


@login_required
def dipendente_turni(request):
    if not (utente_admin(request.user) or utente_dipendente(request.user)):
        messages.error(request, "Accesso non autorizzato.")
        return redirect("logout_view")

    calendario = CalendarioMensile.objects.order_by("-anno", "-mese").first()

    if not calendario:
        messages.error(request, "Nessun calendario disponibile.")
        return redirect("dipendente_dashboard")

    dipendenti = list(Dipendente.objects.filter(attivo=True).order_by("cognome", "nome"))
    dipendente_corrente = Dipendente.objects.filter(utente=request.user).first()

    numero_giorni = pycalendar.monthrange(calendario.anno, calendario.mese)[1]
    giorni = []
    giorni_settimana_it = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]

    for g in range(1, numero_giorni + 1):
        data_corrente = date(calendario.anno, calendario.mese, g)
        giorni.append({
            "numero": g,
            "sigla": giorni_settimana_it[data_corrente.weekday()],
            "weekend": data_corrente.weekday() >= 5,
        })

    assegnazioni = AssegnazioneTurno.objects.filter(
        calendario=calendario
    ).select_related("dipendente")

    mappa = defaultdict(dict)
    riepiloghi = {
        d.id: {"M": 0, "P": 0, "N": 0, "R": 0, "F": 0, "A": 0, "L": 0, "X": 0, "totale": 0}
        for d in dipendenti
    }

    for a in assegnazioni:
        giorno = a.data.day
        mappa[a.dipendente_id][giorno] = a.turno
        riepiloghi[a.dipendente_id][a.turno] += 1
        if a.turno in ["M", "P", "N"]:
            riepiloghi[a.dipendente_id]["totale"] += 1

    righe = []
    for d in dipendenti:
        turni_giornalieri = []
        for g in giorni:
            turno = mappa[d.id].get(g["numero"], "")
            turni_giornalieri.append({
                "valore": turno,
                "weekend": g["weekend"],
            })

        righe.append({
            "dipendente": d,
            "turni": turni_giornalieri,
            "riepilogo": riepiloghi[d.id],
            "evidenziato": dipendente_corrente and d.id == dipendente_corrente.id,
        })

    return render(
        request,
        "scheduler/dipendente_turni.html",
        {
            "calendario": calendario,
            "giorni": giorni,
            "righe": righe,
            "dipendente_corrente": dipendente_corrente,
        }
    )


@login_required
def miei_turni(request):
    if not (utente_admin(request.user) or utente_dipendente(request.user)):
        messages.error(request, "Accesso non autorizzato.")
        return redirect("logout_view")

    calendario = CalendarioMensile.objects.order_by("-anno", "-mese").first()

    if not calendario:
        messages.error(request, "Nessun calendario disponibile.")
        return redirect("dipendente_dashboard")

    dipendente_corrente = Dipendente.objects.filter(utente=request.user).first()

    if not dipendente_corrente:
        messages.error(request, "Nessun dipendente associato al tuo account.")
        return redirect("dipendente_dashboard")

    numero_giorni = pycalendar.monthrange(calendario.anno, calendario.mese)[1]
    giorni = []
    giorni_settimana_it = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]

    for g in range(1, numero_giorni + 1):
        data_corrente = date(calendario.anno, calendario.mese, g)
        giorni.append({
            "numero": g,
            "sigla": giorni_settimana_it[data_corrente.weekday()],
            "weekend": data_corrente.weekday() >= 5,
        })

    assegnazioni = AssegnazioneTurno.objects.filter(
        calendario=calendario,
        dipendente=dipendente_corrente
    )

    mappa = {}
    riepilogo = {"M": 0, "P": 0, "N": 0, "R": 0, "F": 0, "A": 0, "L": 0, "X": 0, "totale": 0}

    for a in assegnazioni:
        giorno = a.data.day
        mappa[giorno] = a.turno
        riepilogo[a.turno] += 1
        if a.turno in ["M", "P", "N"]:
            riepilogo["totale"] += 1

    turni_giornalieri = []
    for g in giorni:
        turno = mappa.get(g["numero"], "")
        turni_giornalieri.append({
            "valore": turno,
            "weekend": g["weekend"],
        })

    return render(
        request,
        "scheduler/miei_turni.html",
        {
            "calendario": calendario,
            "giorni": giorni,
            "dipendente": dipendente_corrente,
            "turni": turni_giornalieri,
            "riepilogo": riepilogo,
        }
    )